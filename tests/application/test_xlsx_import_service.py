"""Tests for XlsxImportService."""
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.application.services.xlsx_import_service import XlsxImportService


class DummyService:
    def __init__(self):
        self.created = []

    async def create_entity(self, **kwargs):
        self.created.append(kwargs)


class DummyEventService(DummyService):
    async def create_event(self, **kwargs):
        self.created.append(kwargs)


@pytest.mark.asyncio
async def test_import_events_tmpdir(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.append(["name", "start_date", "end_date", "characteristics", "backstory"])
    ws.append(["Battle", date(1200, 1, 1), date(1200, 12, 31), "Big fight", "Ancient war"])
    path = tmp_path / "events.xlsx"
    wb.save(path)

    ev_svc = DummyEventService()
    char_svc = DummyService()
    loc_svc = DummyService()
    org_svc = DummyService()
    item_svc = DummyService()
    svc = XlsxImportService(ev_svc, char_svc, loc_svc, org_svc, item_svc)

    result = await svc.import_file("event", path)
    assert result.created == 1
    assert not result.errors
    assert len(ev_svc.created) == 1
    assert ev_svc.created[0]["name"] == "Battle"






def _svc() -> XlsxImportService:
    return XlsxImportService(DummyEventService(), DummyService(), DummyService(), DummyService(), DummyService())


def _write_xlsx(path: Path, headers, rows) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


# ── validate_file ─────────────────────────────────────────────────────────


class TestValidateFile:
    def test_unknown_type(self, tmp_path):
        assert _svc().validate_file("planet", tmp_path / "x.xlsx") == ["Неизвестный тип: planet"]

    def test_type_is_case_insensitive(self, tmp_path):
        errors = _svc().validate_file("EVENT", tmp_path / "x.xlsx")
        assert errors == [f"Файл не найден: {tmp_path / 'x.xlsx'}"]

    def test_missing_file(self, tmp_path):
        errors = _svc().validate_file("event", tmp_path / "nope.xlsx")
        assert len(errors) == 1
        assert "не найден" in errors[0]

    def test_corrupted_file(self, tmp_path):
        p = tmp_path / "bad.xlsx"
        p.write_bytes(b"this is not a zip archive")
        errors = _svc().validate_file("event", p)
        assert len(errors) == 1
        assert "повреждён" in errors[0]

    def test_empty_workbook(self, tmp_path):
        p = tmp_path / "empty.xlsx"
        Workbook().save(p)
        errors = _svc().validate_file("event", p)
        assert errors == ["Файл пустой. Добавьте строку заголовков и данные."]

    def test_missing_required_headers(self, tmp_path):
        p = _write_xlsx(tmp_path / "t.xlsx", ["name"], [["x", "2001-01-01"]])
        errors = _svc().validate_file("character", p)
        assert errors == ["Обязательные столбцы отсутствуют: start_date."]

    def test_valid_file_returns_no_errors(self, tmp_path):
        p = _write_xlsx(tmp_path / "t.xlsx", ["name", "start_date"], [["x", "2001-01-01"]])
        assert _svc().validate_file("item", p) == []


# ── import_file: error branches ───────────────────────────────────────────


class TestImportErrors:
    async def test_unknown_type_raises(self, tmp_path):
        with pytest.raises(ValueError, match="Unsupported entity type"):
            await _svc().import_file("planet", tmp_path / "x.xlsx")

    async def test_missing_file_raises(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            await _svc().import_file("event", tmp_path / "nope.xlsx")

    async def test_validation_failed_returns_result_with_errors(self, tmp_path):
        p = _write_xlsx(tmp_path / "t.xlsx", ["name"], [["x"]])  # start_date missing
        result = await _svc().import_file("event", p)
        assert result.created == 0
        assert result.errors == ["Обязательные столбцы отсутствуют: start_date."]

    async def test_empty_workbook_returns_result(self, tmp_path):
        p = tmp_path / "empty.xlsx"
        # validate_file would flag it as empty → returns validation error result
        Workbook().save(p)
        result = await _svc().import_file("event", p)
        assert result.created == 0
        assert len(result.errors) == 1


# ── import_file: row-level validation ─────────────────────────────────────


class TestRowValidation:
    async def test_row_without_name_is_skipped_with_warning(self, tmp_path):
        p = _write_xlsx(
            tmp_path / "t.xlsx",
            ["name", "start_date"],
            [[None, "2001-01-01"], ["Real", "2001-01-01"]],
        )
        svc = _svc()
        result = await svc.import_file("character", p)
        assert result.created == 1
        assert result.errors == ["Строка 2: пустое имя, пропуск."]
        assert svc._svc_map["character"].created[0]["name"] == "Real"

    async def test_row_without_start_date_is_skipped_with_warning(self, tmp_path):
        p = _write_xlsx(
            tmp_path / "t.xlsx",
            ["name", "start_date"],
            [["NoDate", None], ["Real", "2001-01-01"]],
        )
        svc = _svc()
        result = await svc.import_file("event", p)
        assert result.created == 1
        assert result.errors == ["Строка 2: не задана дата начала, пропуск."]

    async def test_row_exception_is_collected_not_raised(self, tmp_path):
        svc = _svc()
        svc._event_service = DummyEventService()

        async def boom(**kwargs):
            raise RuntimeError("db down")

        svc._event_service.create_event = boom
        p = _write_xlsx(
            tmp_path / "t.xlsx",
            ["name", "start_date"],
            [["A", "2001-01-01"]],
        )
        result = await svc.import_file("event", p)
        assert result.created == 0
        assert len(result.errors) == 1
        assert "RuntimeError" in result.errors[0]
        assert "Строка 2" in result.errors[0]


# ── import_file: dates ────────────────────────────────────────────────────


class TestDates:
    async def test_text_dates_parsed(self, tmp_path):
        p = _write_xlsx(
            tmp_path / "t.xlsx",
            ["name", "start_date", "end_date"],
            [["E", "2001-02-03", "2001-12-31"]],
        )
        svc = _svc()
        result = await svc.import_file("event", p)
        assert result.created == 1
        assert svc._event_service.created[0]["start_date"] == date(2001, 2, 3)
        assert svc._event_service.created[0]["end_date"] == date(2001, 12, 31)

    async def test_native_excel_date_cells(self, tmp_path):
        p = _write_xlsx(
            tmp_path / "t.xlsx",
            ["name", "start_date", "end_date"],
            [["E", date(1200, 3, 4), date(1200, 4, 5)]],
        )
        svc = _svc()
        result = await svc.import_file("character", p)
        assert result.created == 1
        start = svc._svc_map["character"].created[0]["start_date"]
        assert (start.year, start.month, start.day) == (1200, 3, 4)

    async def test_unparseable_date_text_is_treated_as_missing(self, tmp_path):
        p = _write_xlsx(
            tmp_path / "t.xlsx",
            ["name", "start_date"],
            [["E", "not-a-date"]],
        )
        result = await _svc().import_file("event", p)
        assert result.created == 0
        assert result.errors == ["Строка 2: не задана дата начала, пропуск."]

    async def test_event_without_end_date_passes_none(self, tmp_path):
        p = _write_xlsx(tmp_path / "t.xlsx", ["name", "start_date"], [["E", "2001-01-01"]])
        svc = _svc()
        result = await svc.import_file("event", p)
        assert result.created == 1
        assert svc._event_service.created[0]["end_date"] is None

    async def test_entity_without_end_date_defaults_to_start(self, tmp_path):
        p = _write_xlsx(tmp_path / "t.xlsx", ["name", "start_date"], [["E", "2001-01-01"]])
        svc = _svc()
        result = await svc.import_file("location", p)
        assert result.created == 1
        assert svc._svc_map["location"].created[0]["end_date"] == date(2001, 1, 1)


# ── import_file: optional columns ─────────────────────────────────────────


class TestOptionalColumns:
    async def test_character_extras_are_forwarded(self, tmp_path):
        p = _write_xlsx(
            tmp_path / "t.xlsx",
            ["name", "start_date", "personality", "tasks", "music_url"],
            [["C", "2001-01-01", "Bold", "Win", "http://m/1"]],
        )
        svc = _svc()
        await svc.import_file("character", p)
        kw = svc._svc_map["character"].created[0]
        assert kw["personality"] == "Bold"
        assert kw["tasks"] == "Win"
        assert kw["music_url"] == "http://m/1"

    async def test_empty_optional_columns_are_omitted(self, tmp_path):
        p = _write_xlsx(
            tmp_path / "t.xlsx",
            ["name", "start_date", "personality", "tasks", "music_url"],
            [["C", "2001-01-01", None, "", "  "]],
        )
        svc = _svc()
        await svc.import_file("character", p)
        kw = svc._svc_map["character"].created[0]
        # "" cell round-trips as empty (None) → omitted;
        # whitespace-only cell is a real string → forwarded stripped to "".
        assert "personality" not in kw
        assert "tasks" not in kw
        assert kw["music_url"] == ""

    async def test_descriptions_forwarded_to_create(self, tmp_path):
        p = _write_xlsx(
            tmp_path / "t.xlsx",
            ["name", "start_date", "characteristics", "backstory"],
            [["E", "2001-01-01", "Ch", "BS"]],
        )
        svc = _svc()
        await svc.import_file("item", p)
        kw = svc._svc_map["item"].created[0]
        assert kw["characteristics"] == "Ch"
        assert kw["backstory"] == "BS"


# ── import_file: image column ─────────────────────────────────────────────


class TestImageColumn:
    def _png(self, tmp_path: Path, name="art.png") -> Path:
        from PySide6.QtGui import QImage
        p = tmp_path / name
        img = QImage(4, 4, QImage.Format.Format_ARGB32)
        img.fill(0xFF0000FF)
        assert img.save(str(p), "PNG")
        return p

    async def test_relative_image_path_loaded(self, tmp_path):
        self._png(tmp_path)
        p = _write_xlsx(
            tmp_path / "t.xlsx",
            ["name", "start_date", "image"],
            [["C", "2001-01-01", "art.png"]],
        )
        svc = _svc()
        result = await svc.import_file("character", p)
        assert result.created == 1 and not result.errors
        assert svc._svc_map["character"].created[0]["image"].startswith("iVBOR")

    async def test_absolute_image_path_loaded(self, tmp_path):
        img = self._png(tmp_path, "abs.png")
        p = _write_xlsx(
            tmp_path / "t.xlsx",
            ["name", "start_date", "image"],
            [["C", "2001-01-01", str(img)]],
        )
        svc = _svc()
        result = await svc.import_file("organization", p)
        assert result.created == 1 and not result.errors
        assert "image" in svc._svc_map["organization"].created[0]

    async def test_alternate_header_изображение(self, tmp_path):
        self._png(tmp_path)
        p = _write_xlsx(
            tmp_path / "t.xlsx",
            ["name", "start_date", "изображение"],
            [["C", "2001-01-01", "art.png"]],
        )
        svc = _svc()
        result = await svc.import_file("location", p)
        assert result.created == 1 and not result.errors
        assert "image" in svc._svc_map["location"].created[0]

    async def test_missing_image_file_reported(self, tmp_path):
        p = _write_xlsx(
            tmp_path / "t.xlsx",
            ["name", "start_date", "image"],
            [["C", "2001-01-01", "ghost.png"]],
        )
        svc = _svc()
        result = await svc.import_file("character", p)
        assert result.created == 1
        assert result.errors == ["Строка 2: не удалось загрузить изображение «ghost.png»"]
        assert "image" not in svc._svc_map["character"].created[0]

    async def test_unsupported_image_suffix_reported(self, tmp_path):
        (tmp_path / "art.txt").write_text("not an image")
        p = _write_xlsx(
            tmp_path / "t.xlsx",
            ["name", "start_date", "image"],
            [["C", "2001-01-01", "art.txt"]],
        )
        svc = _svc()
        result = await svc.import_file("character", p)
        assert result.created == 1
        assert result.errors == ["Строка 2: не удалось загрузить изображение «art.txt»"]

    async def test_corrupt_image_file_reported(self, tmp_path):
        (tmp_path / "bad.png").write_bytes(b"not a png")
        p = _write_xlsx(
            tmp_path / "t.xlsx",
            ["name", "start_date", "image"],
            [["C", "2001-01-01", "bad.png"]],
        )
        svc = _svc()
        result = await svc.import_file("character", p)
        assert result.created == 1
        assert "не удалось загрузить изображение" in result.errors[0]

    async def test_item_type_ignores_image_column(self, tmp_path):
        self._png(tmp_path)
        p = _write_xlsx(
            tmp_path / "t.xlsx",
            ["name", "start_date", "image"],
            [["I", "2001-01-01", "ghost.png"]],
        )
        result = await _svc().import_file("item", p)
        assert result.created == 1
        assert result.errors == []


# ── import_file: progress ─────────────────────────────────────────────────


class TestProgress:
    async def test_progress_callback_sequence(self, tmp_path):
        p = _write_xlsx(
            tmp_path / "t.xlsx",
            ["name", "start_date"],
            [["A", "2001-01-01"], ["B", "2001-01-01"]],
        )
        calls: list[tuple[int, int]] = []
        result = await _svc().import_file("event", p, progress_callback=lambda c, t: calls.append((c, t)))
        assert calls == [(0, 2), (1, 2), (2, 2)]
        assert result.created == 2

    async def test_progress_callback_optional(self, tmp_path):
        p = _write_xlsx(tmp_path / "t.xlsx", ["name", "start_date"], [["A", "2001-01-01"]])
        result = await _svc().import_file("event", p)
        assert result.created == 1
