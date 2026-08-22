"""Service for importing entities from .xlsx files."""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import load_workbook

from app.application.services.entity_service import EntityService
from app.application.services.event_service import EventService


@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    errors: list[str] | None = None

    def __post_init__(self) -> None:
        if self.errors is None:
            self.errors = []


class XlsxImportService:
    """High-level service for importing events and entities from .xlsx files."""

    def __init__(
        self,
        event_service: EventService,
        character_service: EntityService,
        location_service: EntityService,
        organization_service: EntityService,
        item_service: EntityService,
    ) -> None:
        self._event_service = event_service
        self._svc_map: dict[str, EntityService | EventService] = {
            "event": event_service,
            "character": character_service,
            "location": location_service,
            "organization": organization_service,
            "item": item_service,
        }

    def validate_file(self, entity_type: str, path: str | Path) -> list[str]:
        """Open file, check structure. Returns list of error messages; empty if valid."""
        entity_type = entity_type.lower()
        if entity_type not in self._svc_map:
            return [f"Неизвестный тип: {entity_type}"]
        path = Path(path)
        if not path.exists():
            return [f"Файл не найден: {path}"]
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            return [f"Файл повреждён или не является .xlsx: {e!s}"]
        ws = wb.active
        if ws is None:
            return ["В книге нет активного листа."]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return ["Файл пустой. Добавьте строку заголовков и данные."]
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        required = {
            "event": ["name", "start_date"],
            "character": ["name", "start_date"],
            "location": ["name", "start_date"],
            "organization": ["name", "start_date"],
            "item": ["name", "start_date"],
        }[entity_type]
        missing = [h for h in required if h not in headers]
        if missing:
            return ["Обязательные столбцы отсутствуют: " + ", ".join(missing) + "."]
        return []

    async def import_file(
        self,
        entity_type: str,
        path: str | Path,
        progress_callback: Callable[[int, int], None] | None = None,
    ) -> ImportResult:
        """Import entities of the given type from a .xlsx file."""
        entity_type = entity_type.lower()
        if entity_type not in self._svc_map:
            raise ValueError(f"Unsupported entity type for import: {entity_type}")

        path = Path(path)
        if not path.exists():
            raise FileNotFoundError(path)

        validation_errors = self.validate_file(entity_type, path)
        if validation_errors:
            return ImportResult(created=0, updated=0, errors=validation_errors)

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active

        # Expect the first row to be headers
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return ImportResult(created=0, updated=0, errors=["Файл пустой."])

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        body = rows[1:]

        header_index = {name: idx for idx, name in enumerate(headers)}
        svc = self._svc_map[entity_type]
        result = ImportResult()
        total = len(body)
        for idx, row in enumerate(body, start=2):  # 1-based + header
            if progress_callback:
                progress_callback(idx - 2, total)
            try:
                name = self._get_str(row, header_index, "name")
                if not name:
                    result.errors.append(f"Строка {idx}: пустое имя, пропуск.")
                    continue

                start = self._get_date(row, header_index, "start_date")
                if not start:
                    result.errors.append(f"Строка {idx}: не задана дата начала, пропуск.")
                    continue

                end = self._get_date(row, header_index, "end_date")
                characteristics = self._get_str(row, header_index, "characteristics") or ""
                backstory = self._get_str(row, header_index, "backstory") or ""

                extra: dict[str, Any] = {}
                for col in ("tasks", "personality", "music_url"):
                    val = self._get_str(row, header_index, col)
                    if val is not None:
                        extra[col] = val

                # Image: column "image" or "изображение" — local path (relative to xlsx dir or absolute)
                if entity_type in ("character", "organization", "location"):
                    img_path = (
                        self._get_str(row, header_index, "image")
                        or self._get_str(row, header_index, "изображение")
                    )
                    if img_path:
                        b64 = self._load_image_from_path(path.parent, img_path)
                        if b64:
                            extra["image"] = b64
                        else:
                            result.errors.append(f"Строка {idx}: не удалось загрузить изображение «{img_path}»")

                if entity_type == "event":
                    await self._event_service.create_event(
                        name=name,
                        start_date=start,
                        end_date=end,
                        characteristics=characteristics,
                        backstory=backstory,
                    )
                else:
                    await svc.create_entity(
                        name=name,
                        start_date=start,
                        end_date=end or start,
                        characteristics=characteristics,
                        backstory=backstory,
                        **extra,
                    )
                result.created += 1
            except Exception as exc:  # noqa: BLE001
                result.errors.append(f"Строка {idx}: {exc!r}")
        if progress_callback:
            progress_callback(total, total)
        return result

    def _load_image_from_path(self, base_dir: Path, cell_value: str) -> str | None:
        """Resolve path (relative to base_dir or absolute), load image, return base64 or None."""
        p = Path(cell_value.strip())
        if not p.is_absolute():
            p = (base_dir / p).resolve()
        if not p.exists() or not p.is_file():
            return None
        suffix = p.suffix.lower()
        if suffix not in (".png", ".jpg", ".jpeg", ".bmp", ".gif", ".webp"):
            return None
        try:
            from app.presentation.utils.image_utils import load_and_encode
            return load_and_encode(str(p), max_size=1000)
        except Exception:
            return None

    @staticmethod
    def _get_str(row: Iterable[Any], idx_map: dict[str, int], col: str) -> str | None:
        pos = idx_map.get(col)
        if pos is None or pos >= len(row):
            return None
        val = row[pos]
        if val is None:
            return None
        return str(val).strip()

    @staticmethod
    def _get_date(row: Iterable[Any], idx_map: dict[str, int], col: str) -> date | None:
        pos = idx_map.get(col)
        if pos is None or pos >= len(row):
            return None
        val = row[pos]
        if isinstance(val, date):
            return val
        if val is None:
            return None
        try:
            # Expect YYYY-MM-DD
            parts = str(val).split("-")
            if len(parts) == 3:
                y, m, d = map(int, parts)
                return date(y, m, d)
        except Exception:
            return None
        return None

